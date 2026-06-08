#!/usr/bin/env python3
import argparse
import json
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import re
import sys

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    print(
        "openpyxl is required. Run with Codex bundled Python or install openpyxl in your Python environment.",
        file=sys.stderr,
    )
    raise


MONEY = Decimal("0.000001")


def to_decimal(value):
    if value in (None, ""):
        return Decimal("0")
    return Decimal(str(value)).quantize(MONEY, rounding=ROUND_HALF_UP)


def to_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass
    raise ValueError(f"Unsupported date value: {value!r}")


def date_json(value):
    return value.isoformat() if value else None


def decimal_json(value):
    return format(value.quantize(MONEY, rounding=ROUND_HALF_UP), "f")


def in_period(value, period_start, period_end):
    return value is not None and period_start <= value <= period_end


def eligible_days(entry_date, exit_date, period_start, period_end):
    if not entry_date:
        return 0
    start = max(entry_date, period_start)
    end = min(exit_date or period_end, period_end)
    if start > end:
        return 0
    return (end - start).days + 1


def normalized_operation_type(value):
    text = str(value or "").strip().lower()
    if text in {"продажа", "sale", "sell"}:
        return "sale"
    if text in {"покупка", "purchase", "buy"}:
        return "purchase"
    return text


def investor_key(row_number, external_id, name):
    base = str(external_id or name or f"row-{row_number}").strip().lower()
    base = re.sub(r"\s+", "-", base)
    base = re.sub(r"[^a-zа-я0-9._-]+", "", base)
    return base or f"row-{row_number}"


def parse_workbook(path):
    workbook = load_workbook(path, data_only=True)
    params = workbook["Параметры"]
    operations_sheet = workbook["Операции"]
    expenses_sheet = workbook["Расходы"]
    investors_sheet = workbook["Инвесторы"]

    period_start = to_date(params["B3"].value)
    period_end = to_date(params["B4"].value)
    investor_share_percent = to_decimal(params["B5"].value) * Decimal("100")

    if not period_start or not period_end or period_start > period_end:
        raise ValueError("Invalid quarter dates in Параметры!B3:B4")

    operations = []
    gross_revenue = Decimal("0")
    direct_cost = Decimal("0")

    for row_number in range(4, operations_sheet.max_row + 1):
        operation_date = to_date(operations_sheet.cell(row_number, 1).value)
        operation_type = normalized_operation_type(operations_sheet.cell(row_number, 2).value)
        grams = to_decimal(operations_sheet.cell(row_number, 3).value)
        price = to_decimal(operations_sheet.cell(row_number, 4).value)
        amount = to_decimal(operations_sheet.cell(row_number, 7).value) or (grams * price).quantize(MONEY)

        if not operation_date or not operation_type or amount <= 0:
            continue

        is_inside = in_period(operation_date, period_start, period_end)

        if is_inside and operation_type == "sale":
            gross_revenue += amount
        if is_inside and operation_type == "purchase":
            direct_cost += amount

        operations.append(
            {
                "row": row_number,
                "date": date_json(operation_date),
                "type": operation_type,
                "grams": decimal_json(grams),
                "priceUsdt": decimal_json(price),
                "amountUsdt": decimal_json(amount),
                "insidePeriod": is_inside,
            }
        )

    expenses = []
    operating_expense = Decimal("0")

    for row_number in range(4, expenses_sheet.max_row + 1):
        expense_date = to_date(expenses_sheet.cell(row_number, 1).value)
        amount = to_decimal(expenses_sheet.cell(row_number, 4).value)

        if not expense_date or amount <= 0:
            continue

        is_inside = in_period(expense_date, period_start, period_end)
        if is_inside:
            operating_expense += amount

        expenses.append(
            {
                "row": row_number,
                "date": date_json(expense_date),
                "category": str(expenses_sheet.cell(row_number, 2).value or "").strip(),
                "description": str(expenses_sheet.cell(row_number, 3).value or "").strip(),
                "amountUsdt": decimal_json(amount),
                "insidePeriod": is_inside,
            }
        )

    net_profit = gross_revenue - direct_cost - operating_expense
    investor_pool = Decimal("0")
    if net_profit > 0:
        investor_pool = (net_profit * investor_share_percent / Decimal("100")).quantize(MONEY)

    investors = []
    total_invested = Decimal("0")
    total_weight = Decimal("0")

    for row_number in range(4, investors_sheet.max_row + 1):
        external_id = str(investors_sheet.cell(row_number, 1).value or "").strip()
        name = str(investors_sheet.cell(row_number, 2).value or "").strip()
        entry_date = to_date(investors_sheet.cell(row_number, 3).value)
        exit_date = to_date(investors_sheet.cell(row_number, 4).value)
        amount = to_decimal(investors_sheet.cell(row_number, 5).value)

        if not name or amount <= 0:
            continue

        days = eligible_days(entry_date, exit_date, period_start, period_end)
        weight = (amount * Decimal(days)).quantize(MONEY)
        key = investor_key(row_number, external_id, name)

        total_invested += amount
        total_weight += weight
        investors.append(
            {
                "row": row_number,
                "key": key,
                "externalId": external_id,
                "name": name,
                "entryDate": date_json(entry_date),
                "exitDate": date_json(exit_date),
                "amountUsdt": decimal_json(amount),
                "eligibleDays": days,
                "weight": decimal_json(weight),
                "share": "0",
                "dividendUsdt": "0",
            }
        )

    allocated = Decimal("0")
    payable_indexes = [index for index, investor in enumerate(investors) if Decimal(investor["weight"]) > 0]
    for list_index, investor_index in enumerate(payable_indexes):
        investor = investors[investor_index]
        weight = Decimal(investor["weight"])
        share = Decimal("0") if total_weight <= 0 else (weight / total_weight)
        if investor_pool > 0 and total_weight > 0:
            amount = investor_pool - allocated if list_index == len(payable_indexes) - 1 else (investor_pool * share).quantize(MONEY)
            allocated += amount
        else:
            amount = Decimal("0")
        investor["share"] = str(share)
        investor["dividendUsdt"] = decimal_json(amount)

    return {
        "workbook": str(path),
        "period": {
            "label": f"{period_start.year} Q{((period_start.month - 1) // 3) + 1}",
            "start": date_json(period_start),
            "end": date_json(period_end),
            "investorSharePercent": decimal_json(investor_share_percent),
        },
        "totals": {
            "grossRevenueUsdt": decimal_json(gross_revenue),
            "directCostUsdt": decimal_json(direct_cost),
            "operatingExpenseUsdt": decimal_json(operating_expense),
            "netProfitUsdt": decimal_json(net_profit),
            "investorPoolUsdt": decimal_json(investor_pool),
            "totalInvestedUsdt": decimal_json(total_invested),
            "totalWeight": decimal_json(total_weight),
            "investorCount": len(investors),
            "eligibleInvestorCount": len(payable_indexes),
            "operationCount": len(operations),
            "expenseCount": len(expenses),
        },
        "operations": operations,
        "expenses": expenses,
        "investors": investors,
    }


def main():
    parser = argparse.ArgumentParser(description="Parse Al Amana Gold quarterly workbook into JSON.")
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()
    payload = parse_workbook(args.workbook)
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
